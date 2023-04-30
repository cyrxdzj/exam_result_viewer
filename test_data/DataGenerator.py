import openpyxl
import random
data_name=input("Data name: ")
count=int(input("Data count: "))
workbook = openpyxl.Workbook()
sheet1 = workbook.create_sheet()
sheet1.cell(column=1,row=1).value=data_name
subjects=["语文","数学","英语","道德与法治","历史","生物","地理"]
full_scores=[120,120,90,90,90,100,100]
sheet1.cell(column=1,row=2).value="班级"
sheet1.cell(column=2,row=2).value="姓名"
sheet1.cell(column=3,row=2).value="考号"
for i in range(len(subjects)):
    sheet1.cell(column=i+4,row=2).value=subjects[i]
for i in range(count):
    sheet1.cell(column=1,row=3+i).value=str(i//45+1)
    sheet1.cell(column=2,row=3+i).value="Student %d"%(i+1)
    sheet1.cell(column=3,row=3+i).value="011803%04d"%(i+1)
    for j in range(len(subjects)):
        sheet1.cell(column=j+4,row=3+i).value=random.randint(0,full_scores[j])
workbook.save(data_name+".xlsx")
